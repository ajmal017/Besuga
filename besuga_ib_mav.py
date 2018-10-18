

#ib.connect('127.0.0.1', 7496, clientId=1)

# importem llibreria insync
#from ib_insync import *
#ib = IB()

# definim funcions



def ibConnect(ip,port,clientid):
    # connectem a IBAPI
    ib.connect(ip, port, clientid)

def ibDisconnect():
    ib.disconnect()

#def error(self, reqId: TickerId, errorCode: int, errorString: str):
#    super().error(reqId, errorCode, errorString)
#    print("Error. Id: ", reqId, " Code: ", errorCode, " Msg: ", errorString)

def createWB():
    wb = op.Workbook()
    # set file path
    filepath = r"C:\Users\Usuario\PycharmProjects\Besuga-IB\BesugaIBAPI.xlsx"
    # save workbook
    wb.get_sheet_names()
    sheet =wb.active
    sheet.title
    sheet.title = "Gontech"
    wb.get_sheet_names()
    wb.save(filepath)


def getCurrentDir():
    # get current directory
    cwd = os.getcwd()
    print(cwd)

    # Change directory
    #os.chdir("/path/to/your/folder")

    # List all files and directories in current directory
    #rename files
    #os.rename(old, new)
    print(os.listdir('.'))

def connectIBAPI():
    wb = op.load_workbook("BesugaIBAPI.xlsx")
    sheet = wb.get_sheet_by_name("Gontech")
    ip = sheet["B2"].value
    port = sheet["B3"].value
    clientid = sheet["B4"].value

    print (ip,port,clientid)

    ib.connect(ip, port, clientid)

def importFromExcelOld():
    originalDir = os.getcwd()
    print("originalDir   ", originalDir)

    # get to the right directory
    #os.chdir("G://FinancePortfolioManagemee
    #path = "G://FinancePortfolioManagemeent//aaaBesuga//Python"
    #path = "G:\FinancePortfolioManagemeent\aaaBesuga\Python"
    #os.chdir(path)

    #destinationDir = os.getcwd()
    #print("destinationDir   ", destinationDir)
    # Assign spreadsheet filename to `file`
    excelFileName = "PythonExperiments.xlsx"
    # Load spreadsheet
    xl = pd.ExcelFile(excelFileName)
    # Print the sheet names
    print(xl.sheet_names)
    print(xl._parse_excel)
    # Load a sheet into a DataFrame by name: df1
    dfExcel = xl.parse("Sheet1")
    print(dfExcel)
    #os.chdir("C:\\Users\\Usuario\\PycharmProjects\\Besuga-IB")
def exportToExcel(df,sht):
    # Install `XlsxWriter`
    #pip install XlsxWriter

    # Specify a writer need download XlsxWriter
    writer = pd.ExcelWriter('example.xlsx', engine='xlsxwriter')
    # without downloading XlsxWriter you do not use parameter "engine"
    writer = pd.ExcelWriter('PythonExperiments.xlsx')
    # Write your DataFrame to a file
    df.to_excel(writer, sht)

    # Save the result
    writer.save()

def accountAnalysis():
    accSum =ib.accountSummary()
    accountSummary =[]
    for p in accSum:
        #print(p.tag, p.value)
        accountSummary.append((p.tag,p.value))
        dfaccountSummary= pd.DataFrame(accountSummary)
    print(dfaccountSummary)

    wb = op.load_workbook("BesugaIBAPI.xlsx")
    ws = wb.get_sheet_by_name("AccountData")
    sheet_title = ws.title


    from openpyxl.utils.dataframe import dataframe_to_rows
    rows = dataframe_to_rows(dfaccountSummary)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)


    #ws["A10"].value = dfaccountSummary
    wb.save(r"C:\Users\Usuario\PycharmProjects\Besuga-IB\BesugaIBAPI.xlsx")
    #wb.save("BesugaIBAPI.xlsx")
    #print(reqAccountSummaryTags.AllTags)
def tradeLimitOrder(contract,quantity,ordertype,price):
    global ordresPassadesList
    print("contract: ",contract)
    print("ordertype ", ordertype)
    print("quantity: ", quantity)
    print("price: ", price)

    #order = LimitOrder(ordertype, quantity, price)
    order = LimitOrder(ordertype, quantity, price, tif="GTC",transmit=False)

    ib.qualifyContracts(contract)
    trade = ib.placeOrder(contract, order)
    #ordresPassadesList.append([1])
    #ordresPassadesList.append([contract])

    ib.sleep(1)
    print('TRADE LOG: ', trade.log)

def formatPrice(price):
    precision = 2
    newPrice = np.round(price, precision)
    price = newPrice
    return price

# Anàlisi de posicions obertes
def portfolioAnalysis():
    pfl = ib.portfolio()
    ordresPassadesList = []
    pctProfitList = []

    for p in pfl:
        ib.qualifyContracts(p.contract)
        assert len(ib.reqContractDetails(p.contract)) == 1



        #if p.contract.secType == "OPT":
        if p.contract.secType == "OPT" and p.position < 0:
            posCost = 00
            pctProfitNow = 0

            posCost = p.position * p.averageCost
            if p.position <0:
                pctProfitNow = (1-(p.marketValue / posCost))*100
            else:
                pctProfitNow = ((p.marketValue / posCost)-1) * 100
            print("Symbol: ",p.contract.symbol,"C/P: ",p.contract.right,"Strike: ",p.contract.strike,"Market Price: ", p.marketPrice,"Pos Cost: ",int(posCost),"Pos Value: ",int(p.marketValue),"pctProfitNow: ", int(pctProfitNow))
            pctProfitList.append([p.contract.symbol, p.position, p.contract.right, p.contract.strike, int(pctProfitNow)])
            #print(pctProfitList)
            if pctProfitNow >= pctProfitTarget:
                print("PROFIT TARGET ACHIEVED","Symbol: ", p.contract.symbol, "C/P: ", p.contract.right, "Strike: ", p.contract.strike,"Market Price: ", p.marketPrice,
                      "Pos Cost: ", int(posCost), "Pos Value: ", int(p.marketValue), "pctProfitNow: ",
                      int(pctProfitNow))
                orderType = 'BUY' if p.position < 0 else 'SELL'
                #ask = ib.reqMktData(p.contract).ask
                #bid = ib.reqMktData(p.contract).bid
                #print(p.contract)
                #print(abs(p.position))
                #print(orderType)
                #print(p.marketPrice)

                tickers = ib.reqTickers(p.contract)
                cmdata= ib.reqMktData(p.contract)
                ticker = ib.ticker(p.contract)
                # CONTRACT DETAILS
                #bid = ib.reqMktData(p.contract).bid
                #print("CMDATA.....", cmdata)
                print("TICKER.....", ticker)


                #bid = 0 + cmdata.bid
                #if bid <= 0:
                #   bid = 0
                #modbid = (bid%5)
                #print("modbid....",bid,"   ", modbid)
                #bid = bid + modbid
                #print("bid:    ",bid)
                #ask = 0 + cmdata.ask
                #if ask <=0:
                #   ask = 0
                #modask = (ask % 5)
                #print("modask....", ask, "   ", modask)
                #print("askGreeks    ",ticker.askGreeks)
                #print("lastGreeks    ", ticker.lastGreeks)
                #print("modelGreeks    ", ticker.modelGreeks)

                #nt = ticker.modelGreeks.delta
                #print ("modelgreeks.delta     ",nt)
                #optPrice = ticker.lastGreeks.optPrice
                #print("lastgreeks.optprice", optPrice)

                # COMPUTE EXIT FACTOR
                datetoday = datetime.datetime.now().strftime("%Y%m%d")
                expirationdate = p.contract.lastTradeDateOrContractMonth

                if p.position < 0 and p.contract.secType == "OPT":
                    price = ((p.averageCost* ((100-pctProfitNow))/100))/100
                elif p.position >= 0 and p.contract.secType == "OPT":
                    price = (p.averageCost* (1+(pctProfitNow/100)))/100
                    print ("BBBBBBBBBBBBBBBBBBBBBBBBB    ",p.contract.symbol,"   ", p.averageCost,"   ", p.averageCost* (1+pctProfitNow))
                fmtPrice = formatPrice(price)
                tradeLimitOrder(p.contract, abs(p.position), orderType, fmtPrice)



        if p.contract.symbol == "besugo":


            print("secType ", p.contract.secType)
            print("conId ", p.contract.conId)
            print("symbol ", p.contract.symbol)
            print("lastTradeDateOrContractMonth ", p.contract.lastTradeDateOrContractMonth)
            print("strike ", p.contract.strike)
            print("right ", p.contract.right)
            print("multiplier ", p.contract.multiplier)
            print("exchange ", p.contract.exchange)
            print("primaryExchange ", p.contract.primaryExchange)
            print("currency ", p.contract.currency)
            print("localSymbol ", p.contract.localSymbol)
            print("tradingClass ", p.contract.tradingClass)

            print("position ", p.position)
            print("marketPrice ", p.marketPrice)
            print("marketValue ", p.marketValue)
            print("averageCost ", p.averageCost)
            print("unrealizedPNL ", p.unrealizedPNL)
            print("realizedPNL ", p.realizedPNL)

            print("contract ", p.contract)
            print("account ", p.account)


    #pd.Series(pctProfitList)
    #print("pctProfitList             ", pctProfitList)
    #print("pctProfitList             ", pd.Series(pctProfitList))
    print("pctProfitList             ", pd.DataFrame(pctProfitList),"sheet1")
    exportToExcel(pd.DataFrame(pctProfitList),"sheet1")
    #df = pd.DataFrame(randn(5, 4), index='A B C D E'.split(), columns='W X Y Z'.split())
    #pctProfitListByTicker = pd.groupby(pctProfitList)
    #print("pctProfitList             ", pd.DataFrame(pctProfitList))
    #print("pctProfitListByTicker             ", pd.DataFrame(pctProfitListByTicker))
    #pctOrdersListDF = pd.df.pctProfitList()
    #DF = pd.util.pctProfitList()
    print("ordresPassadesList", ordresPassadesList)
#def initCoses():


#def ultimesRutines():


if __name__ == "__main__":

    # determinem paràmetres gestió portfolio
    pctProfitTarget = 90


    # MAIN BODY

    #importem llibreries
    import numpy as np
    from ib_insync import *
    import openpyxl as op
    import pandas as pd
    import os
    import datetime

    #connectem a IB
    ib = IB()
    # determinem paràmtrees connexio
    #ip= '127.0.0.1'
    #port= 7496
    #clientid=1
    connectIBAPI()

    #getCurrentDir()
    #createWB()

    #importFromExcel()
    # inicialització de variables etc
    #initCoses()

    # analitzem Accoount
    accountAnalysis()

    # analitzem posicions obertes
    portfolioAnalysis()

    # fem últimes rutines
    #ultimesRutines()



    # desconnectem de IBAPI
    ibDisconnect()

